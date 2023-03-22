# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


"""

pip install kiwipiepy
https://bab2min.github.io/kiwipiepy/v0.14.1/kr/

"""
from collections import Counter
import re
# from nltk.stem import WordNetLemmatizer
# from nltk.tag import pos_tag
# from nltk.tokenize import word_tokenize
import pandas as pd
import codecs
from openpyxl import load_workbook
from gensim import corpora
import gensim
from kiwipiepy import Kiwi
kiwi = Kiwi()
brands = {}
unbrands = {}
stopwords = []

"""
fashion domain specific stopwords
FDSS = sigma N intersection trendreport_y_s_topK 


"""
def brand_mapping(readData):
    text = readData.lower()
    for key in brands.keys():
        text = text.replace(key, brands[key])

    return text
def brand_unmapping(readData):
    text = readData.lower()
    for key in brands.keys():
        text = text.replace(brands[key], key)

    return text
def cleanText(readData):
    #text = re.sub('[=+,#/\\\\?:;©^$.@*\"※~&ㆍ!』\\‘|\(\)\[\]\<\>`\'》]', '', readData)
    #text = re.sub('\d\d\d\d', '', text)
    # 소문자 작업
    text = readData.lower()
    # 브랜드 명 다른 이름으로 맵핑
    text = brand_mapping(text)
    text = re.sub('[=+,#/\\\\?:;©^$@*\"※~&ㆍ!』\\‘|\(\)\[\]\<\>`\'》]', '', text)
    
    return text

def writeKeywordScore(get_cells):
    """
    :param get_cells: 
    :return:
    전처리 과정 필요
    1. stopword 정의
    2. 형태소분석
    """
    words = []
    for row in get_cells:
        for cell in row:
            # print(cell.value)
            content = cell.value
            content = cleanText(content)
            lines = content.strip().split('\n')
            pre_word = ""
            pre_pre_word = ""

            for line in lines:
                ws = line.strip().split(' ')
                for w in ws:
                    if len(w) > 0:
                        words.append(w)
                    if pre_word == "":
                        pre_word = w
                        continue
                    words.append(pre_word + " " + w)
                    if pre_pre_word == "":
                        pre_pre_word = pre_word
                        continue
                    words.append(pre_pre_word + " " + pre_word + " " + w)
                    pre_pre_word = pre_word
                    pre_word = w
    word_count_dict = Counter(words)
    word_count_dict = sorted(word_count_dict.items(), key=(lambda x: x[1]), reverse=True)
    tmp = 0
    write_topicScore = codecs.open("keywordScore_all.txt", 'w', 'utf-8')
    for value in word_count_dict:
        if value[1] > 2:
            # if "ï¿½" in value[0]:
            #     continue
            write_topicScore.write(value[0] + "\t" + str(value[1]) + "\n")
            write_topicScore.flush()
            print(value[0] + "\t" + str(value[1]))
            tmp += 1
    write_topicScore.close()
# Press the green button in the gutter to run the script.

def writeLDA(get_cells):
    NUM_TOPICS = [10]
    num_words = 10
    thresholds = ['0.05']
    documents = []
    for row in get_cells:
        for cell in row:
            # print(cell.value)
            content = cell.value
            content = cleanText(content.strip())
            lines = content.strip().split('\n')
            words = []
            for line in lines:
                ws = line.strip().split(' ')
                for w in ws:
                    if len(w) > 0:
                        words.append(w)
            documents.append(words)
    pd.DataFrame({'document':documents})
    dictionary = corpora.Dictionary(documents)
    corpus = [dictionary.doc2bow(text) for text in documents]
    for NUM_TOPIC in NUM_TOPICS:
        ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=NUM_TOPIC, id2word=dictionary, passes=15)
        # presults = ldamodel.print_topics(num_words=num_words,num_topics=NUM_TOPICS)
        results = ldamodel.show_topics(num_topics=NUM_TOPIC, num_words=num_words, formatted=False)
        lda_writer = codecs.open("lda_all.txt", 'w', 'utf-8')
        for tid, twords in results:
            lda_writer.write("{}".format(tid))
            for tword, score in twords:
                lda_writer.write("|{},{}".format(tword, score))
            lda_writer.write('\n')
        lda_writer.close()

def writeKeywordScore_season(key, list):
    words = []
    for content in list:
        content = cleanText(content)
        lines = content.strip().split('\n')
        pre_word = ""
        pre_pre_word = ""

        for line in lines:
            # ws = line.strip().split(' ')
            ws = kiwi.tokenize(line.strip())
            for token in ws:
                if token.tag[0] is not 'N':
                    continue
                w = token.form


                if len(w) > 0 and w not in stopwords:
                    words.append(w)
                if pre_word == "":
                    pre_word = w
                    continue
                words.append(pre_word + " " + w)
                if pre_pre_word == "":
                    pre_pre_word = pre_word
                    continue
                words.append(pre_pre_word + " " + pre_word + " " + w)
                pre_pre_word = pre_word
                pre_word = w
    word_count_dict = Counter(words)
    word_count_dict = sorted(word_count_dict.items(), key=(lambda x: x[1]), reverse=True)
    tmp = 0
    key = re.sub('[=+,#/\\\\?:;©^$@*\"※~&ㆍ!』\\‘|\(\)\[\]\<\>`\'》]', '', key)
    write_topicScore = codecs.open("keywordScore_"+key+".txt", 'w', 'utf-8')
    for value in word_count_dict:
        if value[1] > 2:
            # if "ï¿½" in value[0]:
            #     continue
            item = value[0]
            if item in unbrands:
                item = unbrands[item]
            write_topicScore.write(item + "\t" + str(value[1]) + "\n")
            write_topicScore.flush()
            # print(value[0] + "\t" + str(value[1]))
            tmp += 1
    write_topicScore.close()

def writeLDA_season(key, list):
    NUM_TOPICS = [10]
    num_words = 10
    thresholds = ['0.05']
    documents = []
    for content in list:
        content = cleanText(content.strip())
        lines = content.strip().split('\n')
        words = []
        for line in lines:
            # ws = line.strip().split(' ')
            ws = kiwi.tokenize(line.strip())

            for token in ws:
                w = token.form

                if token.tag[0] is 'N' and len(w) > 0 and w not in stopwords:
                    words.append(w)
        documents.append(words)
    pd.DataFrame({'document':documents})
    dictionary = corpora.Dictionary(documents)
    corpus = [dictionary.doc2bow(text) for text in documents]
    for NUM_TOPIC in NUM_TOPICS:
        ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=NUM_TOPIC, id2word=dictionary, passes=15)
        # presults = ldamodel.print_topics(num_words=num_words,num_topics=NUM_TOPICS)
        results = ldamodel.show_topics(num_topics=NUM_TOPIC, num_words=num_words, formatted=False)
        key = re.sub('[=+,#/\\\\?:;©^$@*\"※~&ㆍ!』\\‘|\(\)\[\]\<\>`\'》]', '', key)
        lda_writer = codecs.open("lda_"+key+"_"+str(NUM_TOPICS)+"_"+str(num_words)+".txt", 'w', 'utf-8')
        for tid, twords in results:
            lda_writer.write("{}".format(tid))
            # print("{}".format(tid), end="")
            for tword, score in twords:
                item = tword
                if item in unbrands:
                    item = unbrands[item]
                lda_writer.write("|{},{}".format(item, score))
                # print("|{},{}".format(item, score), end="")
            lda_writer.write('\n')
            # print('\n', end="")
        lda_writer.close()
#   오늘 백화점에서  asdasd asdasd No.21 을 샀다.
#  오늘 백화점 No.21

if __name__ == '__main__':
    print("start")
    # result = kiwi.tokenize('테스트입니다.')
    # for token in result:
    #     print(f"{token.form}\t{token.tag}")

    load_wb = load_workbook("report.xlsx")
    load_ws = load_wb['Sheet1']
    get_F_cells = load_ws['F2': 'F317']
    get_I_cells = load_ws['I2': 'I317']

    # for brand name
    brand_wb = load_workbook("brand_list.xlsx")
    brand_ws = brand_wb['Sheet1']
    get_A_brand = brand_ws['A2':'A258']


    # for brand name
    for index in range(len(get_A_brand)):
        brand = get_A_brand[index][0].value
        brand = brand.lower()
        brands[brand] = "brand_"+str(index)
        unbrands["brand_"+str(index)] = brand
        kiwi.add_user_word(brands[brand], "NNP")
        # print(get_A_brand[index][0].value)


    # for f_stopwords
    f_stopwords = load_workbook("f_stopwords.xlsx")
    stopwords_ws = f_stopwords['Sheet1']
    get_A_stopwords = stopwords_ws['A2':'A139']
    for index in range(len(get_A_stopwords)):
        s_words = get_A_stopwords[index][0].value
        print(s_words)
        s_words = s_words.lower()
        stopwords.append(s_words)


    season_dict = dict()
    for index in range(len(get_F_cells)):
        season = get_F_cells[index][0].value
        if get_F_cells[index][0].value in season_dict:
            content_list = season_dict[season]
            content_list.append(get_I_cells[index][0].value)
        else:
            content_list = []
            season_dict[season] = content_list
            content_list.append(get_I_cells[index][0].value)
    for key in season_dict:
        content_list = season_dict[key]
        writeKeywordScore_season(key, content_list)
        writeLDA_season(key, content_list)
    get_cells = load_ws['I2': 'I3']
    writeKeywordScore(get_cells)
    writeLDA(get_cells)
